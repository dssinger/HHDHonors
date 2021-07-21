#!/usr/bin/env python3
# vim: set fileencoding=utf-8 :
""" Create the honors.csv file for the mail merge.
  For each honor (unique to service, of course), find all those sharing
  the honor.  If they share an address, combine them.
  
  Use the Honor ID to make matches.
  
  Generate a new spreadsheet with one line per person/couple per honor,
  with a new field for "sharer".

"""
import sys

from openpyxl import load_workbook
import csv
import datetime
import re
import copy

datemode = 1  # make it global


def space(s):
    return ' '.join(s.split())


def normalizeService(service):
    """ Convert a service into full form, removing any sequence numbers and 'Service" """
    if '-' in service:
        service = service.split('-', 1)[1]
    service = service.strip()
    service = service.replace('YK', 'Yom Kippur')
    service = service.replace('RH', 'Rosh Hashanah')
    service = service.replace(' Service', '')
    return service


def normalizeHonor(honor):
    """Take the full description (or short description) of an honor and:
  1)  make it lower-case
  2)  remove all non-alphamerics (including spaces)
  """
    return ''.join(re.split(r'\W+', honor)).lower()


def stringify(value):
    """ Convert values to strings """
    # Let's normalize everything to strings/unicode strings
    if isinstance(value, (int, float)):
        value = '%d' % value
    if isinstance(value, bool):
        value = '1' if value else '0'
    elif isinstance(value, (datetime.datetime, datetime.date)):
        value = ('%s' % value)[0:10]

    return value


class Service:
    services = {}
    labels = {}

    @classmethod
    def setlabels(self, labels):
        self.labels = labels

    @classmethod
    def find(self, key):
        return self.services[normalizeService(key)]

    def __init__(self, row):
        # Convert the items in the row into attributes of the object
        for x in range(len(row)):
            attrname = self.labels[x]
            if attrname not in ('time', 'arrive', 'date', 'early'):
                row[x] = stringify(row[x])
            self.__dict__[attrname] = row[x]

        # Normalize the service name
        self.service = normalizeService(self.service)
        self.daypart = normalizeService(self.daypart)

        # And make "early" an integer
        self.early = int(self.early)

        # Now, let's get the times and dates into pretty form and figure out if this service is on Shabbat or not.
        start = datetime.datetime.combine(self.date, self.time)
        arrive = start - datetime.timedelta(minutes=self.early)

        if start.weekday() == 4 and start.hour >= 19:
            self.isShabbat = True
        elif start.weekday() == 5 and start.hour <= 19:
            self.isShabbat = True
        else:
            self.isShabbat = False

        self.time = start.strftime('%I:%M %p').lstrip('0')
        self.arrive = arrive.strftime('%I:%M %p').lstrip('0')
        self.date = start.strftime('%A, %B %d, %Y').replace(' 0', ' ')

        self.services[self.service] = self


class Honor:
    honors = {}
    labels = {}
    all = []

    @classmethod
    def register(self, honoree):
        key = honoree.honor
        if key not in self.honors:
            self.honors[key] = Honor(honoree.honor, honoree.service)
        self.honors[key].sharers.append(honoree)
        self.honors[key]
        return self.honors[key]

    @classmethod
    def setlabels(self, labels):
        self.labels = labels

    @classmethod
    def find(self, honorid):
        return self.honors.get(honorid, None)

    def assign(self, person):
        try:
            self.sharers[person.household_id].addname(person)
        except KeyError:
            self.sharers[person.household_id] = Honoree(person)

    def __init__(self, row):
        # This method defines an honor and puts it in the class indexed by the HonorID+alternative
        # the service has the sequence number removed, and RH and YK expanded
        # print(row)
        row.service = normalizeService(row.service)

        # Convert the items in the row into attributes of the object.  Make some substitutions in attribute name.
        for x in self.labels:
            attrname = x
            if attrname == 'from':
                attrname = 'fromtext'
            elif attrname == 'to':
                attrname = 'totext'
            self.__dict__[attrname] = stringify(row.__dict__[x])
        self.sharers = {}  # Indexed by membership ID
        self.honorid = self.honorid + self.alternative.strip()
        self.honors[self.honorid] = self

        # See if there's a file associated with the service; if so, do Shabbat adjustment if need be.
        if self.cuesheet.lower() in ('y', 's', 'o'):
            if Service.find(self.service).isShabbat and self.cuesheet.lower() == 's':
                self.filename = self.honorid + 's.pdf'
            else:
                self.filename = self.honorid + '.pdf'
        else:
            self.filename = ''

        # And finally, append to the list of honors in order of appearance
        self.all.append(self)

    def __repr__(self):
        ans = "{'sharers': '%s'," % self.sharers
        llist = []
        for p in self.labels:
            if p == 'to' or p == 'from':
                p = p+'text'
            llist.append(p)
        ans += ', '.join(["'%s': '%s'" % (x, getattr(self, x, '***')) for x in llist])
        return ans

    def sharing(self):
        """ Return a list, with a dict in each position:
      me: me
      them: a comma separated list of other sharers' names, with 'and' at the end
    """
        res = []
        sharers = [self.sharers[k] for k in self.sharers]
        #sharers.sort(key=lambda x: x.dear)
        for i in range(len(sharers)):
            entry = {}
            entry['me'] = sharers[i]
            them = [sharers[j].fullnames for j in range(len(sharers)) if j != i]
            if len(them) == 0:
                entry['them'] = ''
            elif len(them) == 1:
                entry['them'] = them[0]
            elif len(them) == 2:
                entry['them'] = ' and '.join(them)
            else:
                them[-1] = 'and ' + them[-1]
                entry['them'] = ', '.join(them)
            res.append(entry)
        return res


class DividedReading:
    all = {}

    def __init__(self, row):
        self.part = row.part
        self.page = row.page
        self.beginswith = row.beginswith
        self.endswith = row.endswith
        self.cue = row.cue
        if self.part not in self.all:
            self.all[self.part] = [self]
        else:
            self.all[self.part].append(self)


class Honoree:
    def __init__(self, person):
        self.firstname = person.firstname.strip()
        self.lastname = person.lastname.strip()
        self.nickname = person.nickname.strip()
        if person.title.strip() == 'Rabbi':
            self.title = 'Rabbi '
        elif person.title.strip() == 'Cantor':
            self.title = 'Cantor '
        else:
            self.title = ''
        self.fullname = self.title + person.displayname
        self.fullnames = self.fullname
        self.addr = person.address
        if person.address2:
            self.addr += '\n' + person.address2
        self.city = person.city
        self.state = person.state
        self.zip = person.zip
        self.csz = self.city + ' ' + self.state + ', ' + self.zip
        self.address = '\n'.join([self.addr, self.csz])
        self.dear = self.firstname
        self.sendpaper = False
        self.emails = []
        if person.email:
            self.emails.append(person.email)

    def addname(self, person):
        self.firstname = [self.firstname, person.firstname]
        self.dear = ' and '.join(self.firstname)
        if self.lastname != person.lastname:
            self.lastname = [self.lastname, person.lastname]
            self.fullname = [self.fullname, person.firstname + ' ' + person.lastname]
            self.fullnames = ' and '.join(self.fullname)
        else:
            self.fullnames = ' and '.join(self.firstname) + ' ' + self.lastname
        if person.email and person.email not in self.emails:
            self.emails.append(person.email)


def getLabelsFromSheet(sheet):
    """Returns all of the labels from a spreadsheet as a dict"""
    labels = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    labels = [''.join(p.split()).lower() for p in labels]
    ret = dict(list(zip(labels, list(range(len(labels))))))
    # Provide two-way associativity
    for p in list(ret.keys()):
        ret[ret[p]] = p
    return ret

class Student:
    """ Contains information about students with honors """
    students = {}
    def __init__(self, row):
        row.extend(('', '', '', '', ''))  # Ensure enough values
        (self.firstname, self.lastname, self.parent1name, self.parent2name, self.email) = row[0:5]
        self.parent1 = People.findbyname(self.parent1name)
        self.parent2 = People.findbyname(self.parent2name)
        self.key = self.firstname + ' ' + self.lastname
        self.students[self.key] = self

    @classmethod
    def findbyname(cls, name):
        return cls.students.get(name, None)



if __name__ == '__main__':

    import os
    from gsheet import GSheet
    from parms import Parms

    parms = Parms()
    os.chdir(parms.datadir)

    # We need to load service information first, because that determines whether we use Shabbat files or regular ones.
    # The "Services Master.xls" file has information about each service, which we use in preference to that in the HHD Honors file.
    services = load_workbook(parms.services, data_only=True)
    sheet = services.active
    # Put the labels into the Service class
    Service.setlabels(getLabelsFromSheet(sheet))

    # Now, load the services into the class
    for inrow in sheet.iter_rows(min_row=2, values_only=True):
        Service(list(inrow))

    # Now, load membership data
    from people import People
    People.loadpeople(parms.roster)

    # Load students
    if parms.students:
        import csv
        with open(parms.students, 'r') as csvfile:
            srdr = csv.reader(csvfile)
            next(srdr)  # Throw away the headers
            for row in srdr:
                Student(row)





    # Load all possible honors from the master file

    master = GSheet(parms.honorsmaster, parms.apikey)
    Honor.setlabels(master.labels)

    for row in master:
        Honor(row)
    # Honor.all has all the honors.

    # Process divided readings if we have them this year:
    sheetname = getattr(parms, 'dividedsheetname', '')
    if sheetname:
        dr = GSheet(parms.honorsmaster, parms.apikey, sheetname=sheetname)
        for row in dr:
            DividedReading(row)

    # Now, process the assignment file:
    sheetname = getattr(parms, 'assignmentsheetname', '')
    assignments = GSheet(parms.assignments, parms.apikey, sheetname)
    for row in assignments:
        if not row.honorid:
            continue  # No honorid for many office-managed roles.
        honor = Honor.find(row.honorid + row.alternative)
        if not honor:
            print('Could not find honor', row.honorid)
            continue
        for name in (row.name1, row.name2, row.name3, row.name4, row.name5, row.name6, row.name7, row.name8):
            if name:
                name = ' '.join(name.strip().split())

                # Strip 'and team'
                try:
                    teamptr = name.index(' and team')
                    name = name[:teamptr]
                except ValueError:
                    pass

                lname = name.lower()
                skipme = False
                for item in ('xxx', 'anniversary', 'confirmation', 'conf', '35+', 'teens', 'birthdays', 'photo'):
                    skipme = skipme or item in lname.split()
                if not skipme:
                    person = People.findbyname(name)
                    if person:
                        honor.assign(person)
                    else:
                        student = Student.findbyname(name)
                        if student:
                            p = copy.copy(student.parent1)
                            p.firstname = student.firstname
                            p.lastname = student.lastname
                            p.nickname = student.firstname
                            p.key = student.key
                            p.displayname = student.key
                            honor.assign(p)
                            if student.parent2:
                                honor.sharers[p.household_id].emails.append(student.parent2.email)
                        else:
                            print('Could not find %s for honor %s' % (name, row.honorid))
                else:
                    print(f'Honor {row.honorid} ({row.description}) for {name} must be handled by the office.')

    ## OK, now we can create the updated spreadsheet
    ##   and build the batch file to print the cue sheets

    with open(parms.honorscsv,'w') as outfile:
        outcsv = csv.writer(outfile)
        outcsv.writerow(('Dear',
                          'Full_Name',
                          'Family_Address',
                          'Family_CSZ',
                          'Email1',
                          'Email2',
                          'Send_Paper',
                          'Service',
                          'Service_Date',
                          'Service_Time',
                          'Location',
                         'Early',
                         'Arrive',
                         'Holiday',
                         'Rabbi',
                         'HonorID',
                         'Honor',
                         'Book',
                         'Pages',
                         'Cue',
                         'FromText',
                         'ToText',
                         'Filename',
                         'Sharing',
                         'Explanation',
                         'Adminname',
                         'Adminemail',
                         'President'
                         ))

        cuesheets = {}
        gscmd = "gs -sDEVICE=pdfwrite -dPDFSETTINGS=/default -dNOPAUSE -dQUIET -dBATCH -sOutputFile=-"
        for theHonor in Honor.all:
            #print(f'{theHonor.honorid} ({theHonor.honorforletter}): {theHonor.sharing()}')
            if theHonor.honorforletter.lower() == 'none':
                continue  # Skip honors that don't need letters generated.

            if not theHonor.sharing():
                continue  # Skip honors with no assignees

            # Figure out page and book.
            if theHonor.pagestart:
                theHonor.book = 'Mishkan Hanefesh'
                if theHonor.pageend:
                    pages = 'pages %s-%s' % (theHonor.pagestart, theHonor.pageend)
                else:
                    pages = 'page %s' % (theHonor.pagestart)
            else:
                theHonor.book = ''
                pages = ''

            theService = Service.services[theHonor.service]

            honorname = theHonor.honorforletter.strip()
            if not honorname:
                honorname = theHonor.description

            for (num, s) in enumerate(theHonor.sharing()):
                try:
                    email1 = s['me'].emails[0]
                except IndexError:
                    email1 = ''
                try:
                    email2 = s['me'].emails[1]
                except:
                    email2 = ''
                if '@' not in email1 or (email2 and '@' not in email2):
                    print('Missing email for %s' % s['me'].fullnames)
                    s['me'].sendpaper = True

                # Handle divided readings:
                if len(theHonor.sharing()) > 1:
                    if theHonor.honorid in DividedReading.all:
                        dr = DividedReading.all[theHonor.honorid][num]
                        pages = 'page ' + dr.page
                        cue = dr.cue
                        fromtext = dr.beginswith
                        totext = dr.endswith
                    else:
                        cue = theHonor.cue
                        fromtext = theHonor.fromtext
                        totext = theHonor.totext
                    if theHonor.filename:
                        fparts = theHonor.filename.split('.')
                        fparts[0] = f'{fparts[0]}-{num + 1}'
                        filename = '.'.join(fparts)
                        if not os.path.exists(os.path.join(parms.cuedir, filename)):
                            filename = theHonor.filename
                            if not os.path.exists(os.path.join(parms.cuedir, filename)):
                                print(f"Cannot find cuesheet for {filename} or {'.'.join(fparts)}")
                    else:
                        filename = ''
                    honorid = f'{theHonor.honorid}-{num + 1}'
                else:
                    cue = theHonor.cue
                    fromtext = theHonor.fromtext
                    totext = theHonor.totext
                    filename = theHonor.filename
                    honorid = theHonor.honorid

                outcsv.writerow((s['me'].dear,
                                 s['me'].fullnames,
                                 s['me'].addr,
                                 s['me'].csz,
                                 email1,
                                 email2,
                                 s['me'].sendpaper,
                                 theService.service,
                                 theService.date,
                                 theService.time,
                                 theService.location,
                                 theService.early,
                                 theService.arrive,
                                 theService.daypart,
                                 theService.rabbi,
                                 honorid,
                                 honorname.replace('\\u2026', '...'),
                                 theHonor.book,
                                 pages,
                                 cue.replace('\\u2026', '...'),
                                 fromtext,
                                 totext,
                                 filename,
                                 s['them'],
                                 theHonor.explanation.replace('\\u2026', '...'),
                                 parms.adminname,
                                 parms.adminemail,
                                 parms.president))

                if theHonor.filename:
                    snum = theHonor.filename[0]
                    if snum not in cuesheets:
                        cuesheets[snum] = []
                    cuesheets[snum].append((theHonor.filename, honorname, s['me'].fullnames))


    # Create the honors database
    if parms.honorsdb:
        try:
            os.remove(parms.honorsdb)
        except FileNotFoundError:
            pass
        import subprocess
        subprocess.run(('sqlite3', parms.honorsdb), input=f'.mode csv\n.import {parms.honorscsv} honors\n',
                               text=True)



    # Finally, write the printing batch files
    import codecs

    for k in cuesheets:
        batfile = codecs.open('print%s.sh' % k, 'w', encoding='utf-8')
        batfile.write('#!/bin/bash\n')
        batfile.write('cd "../New HHD Cues and Readings"\n')
        batfile.write('gs -dBATCH -dNOPAUSE -q -sDEVICE=pdfwrite -sOutputFile=%s.pdf \\\n' % k)
        # cuesheets[k].reverse()
        for s in cuesheets[k]:
            batfile.write('    %s\\\n' % s[0])

        batfile.write('\n')
        for s in cuesheets[k]:
            batfile.write('# %s - %s - %s\n' % s)
        batfile.close()
