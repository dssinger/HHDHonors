#!/usr/bin/env python3
"""People (for HHD, based on ShulCloud data)

Creates an array of people indexed by "firstname lastname", containing their address fields.
If any names are duplicated, complains.

ShulCloud has ONE record per household, containing all adults in the household.

"""

from openpyxl import load_workbook
import re
from datetime import datetime, date



def getLabelsFromSheet(sheet):
    """Returns all of the labels from a spreadsheet as a dict
     Labels are normalized by converting them to lower case,
        removing any leading "home-",
        removing any "'s"
        replacing any '_' with spaces,
        removing any non-alphamerics (including spaces),
        replacing spaces with '_'.
        We treat 'first_name', 'last_name', and 'id' specially to avoid problems elsewhere in the code.
        """
    labels = []
    for p in [cell.value for cell in sheet[1]]:
        p = p.lower()
        if p.startswith('home-'):
            p = p[5:]
        p = p.replace("'s", "")
        p = p.replace('_', ' ')
        p = '_'.join(re.split(r'\W+', p))
        if p.startswith('_'):
            p = p[1:]
        if p.endswith('_'):
            p = p[:-1]
        p = p.replace('first_name', 'firstname')
        p = p.replace('last_name', 'lastname')
        if p == 'id':
            p = 'household_id'
        labels.append(p)  
    ret = dict(list(zip(labels, list(range(len(labels))))))
    # Provide two-way associativity
    for p in list(ret.keys()):
      ret[ret[p]] = p
    return ret

def stringify(value):
    """ Convert values to strings """
    # Let's normalize everything to strings/unicode strings
    if isinstance(value, (int, float)):
        value = '%d' % value
    if isinstance(value, bool):
        value = '1' if value else '0'
    elif isinstance(value, (datetime, date)):
        value = ('%s' % value)[0:10]

    return str(value).strip()

def normalize(value):
    # Convert to string, and get rid of extraneous spaces
    return ' '.join(stringify(value).split())
    
class Nickname:
    nicknames = {}
    def __init__(self, first, last, newfirst='', newlast=''):
        self.first = first
        self.last = last
        self.newfirst = newfirst if newfirst else first
        self.newlast = newlast if newlast else last
        self.key = self.first + ' ' + self.last
        self.newkey = self.newfirst + ' ' + self.newlast
        self.nicknames[self.key] = self
        
    @classmethod
    def find(self, key):
        try:
            return self.nicknames[key]
        except KeyError:
            return None

        
    @classmethod
    def setnicknames(self):
        return
        Nickname('Ivan (Rusty)', 'Gralnik', 'Rusty')
        Nickname('S. Henry', 'Stern', 'Henry')
        Nickname('Itzhak', 'Nir', 'Itzik')
        Nickname('Isabelle', 'Schneider', 'Billee')
        Nickname('Rebecca Katz', 'Tedesco', 'Rebecca', 'Katz Tedesco')
        Nickname('Hugh', 'Seid-Valencia', 'Rabbi Hugh', 'Seid-Valencia')
        Nickname('Julia', 'Hartman', 'Julie')
        Nickname('Marilyn', 'Keelan', 'Lindy')
        Nickname('Michael', 'Adelman', 'Mickey')
        Nickname('Jeffrey', 'Segol', 'Jeff')
        Nickname('Adrian E', 'Cerda', 'Adrian')
        Nickname('Stephen', 'Jackson', 'Steve')
        Nickname('Sara', 'Sperling-Mintz', 'Sara', 'Mintz')
        Nickname('Barry and Lisa', 'Cheskin', 'Barry', 'Cheskin')
        Nickname('Gretchen', 'Sand-Preville', 'Gretchen', 'Preville')
        
        
    def __repr__(self):
        return('key: %s, first: %s => %s, last: %s => %s' % (self.key, self.first, self.newfirst, self.last, self.newlast))


class People:
    linenumber = 1
    dummycount = 0
    people = {}
    debug = False
    synonyms = {
                'dr': 'Drive',
                'st': 'Street',
                'ave': 'Avenue',
                'rd': 'Road',
                'ct': 'Court',
                'blvd': 'Boulevard',
                'av': 'Avenue',
                'ln': 'Lane',
                'wy': 'Way',
                'cir': 'Circle',
                'n': 'North',
                'no': 'North',
                'PO': 'P.O.',
                'pl': 'Place',
                's': 'South',
                'w': 'West',
                'e': 'East'
            }
            
    namesyns = (('Bob', 'Robert', 'Robbie', 'Rob'),
                ('Margie', 'Marjorie'),
                ('Joe', 'Joseph'),
                ('Irv', 'Irving'),
                ('Andy', 'Andrew'),
                ('Rich', 'Richard'),
                ('Josh', 'Joshua'),
                ('Cindy', 'Cynthia'),
                ('Kim', 'Kimberly'),
                ('Ed', 'Edward'),
                ('Nick', 'Nicholas'),
                ('Debbie', 'Deborah', 'Deb', 'Debra'),
                ('Zach', 'Zack', 'Zachary'),
                ('Rochelle', 'Shell'),
                ('Barb', 'Barbara', 'Barbra'),
                ('Liz', 'Elizabeth', 'Elisabeth'),
                ('Mickey', 'Mike', 'Michael'),
                ('John', 'Jon', 'Jonathan'),
                ('Mady', 'Madeleine'),
                ('Pat', 'Patricia')
                )
    namesyns = ()
                
    Nickname.setnicknames()
                
    @classmethod
    def setlabels(self, labels):
        self.labels = labels

    @classmethod
    def getlinenumber(self):
        self.linenumber += 1
        return self.linenumber

    @classmethod
    def loadpeople2(self, fn, debug=False):   # This version is for years with two adults per record
        self.debug = debug
        commonfields = ('household_id', 'address', 'address2', 'city', 'state', 'zip' )
        personfields = ('email', 'title', 'firstname', 'lastname', 'nickname')
        firstnamecol = personfields.index('firstname')
        nicknamecol = personfields.index('nickname')
        emailcol = personfields.index('email')
        thefields = personfields + commonfields
        p1fields = ['primary_' + f for f in personfields] + list(commonfields)
        p2fields = ['secondary_' + f for f in personfields] + list(commonfields)
        s = load_workbook(fn)
        sheet = s.active
        self.setlabels(getLabelsFromSheet(sheet))
        for inrow in sheet.iter_rows(min_row=2):
            # Clean up the values and put them in a dictionary indexed by column label
            values = [normalize(v.value) for v in inrow]
            row = {self.labels[i]: values[i] for i in range(len(values))}


            # if only the primary in a row has an email, give it to the secondary, too
            if not row['secondary_email']:
                row['secondary_email'] = row['primary_email']

            for fgroup in (p1fields, p2fields):
                # Load each person, taking firstname synonyms into account
                altnames = False
                # Handle firstname synonyms:
                pinfo = [row[field] for field in fgroup]
                People(pinfo, thefields)
                if  pinfo[nicknamecol] and pinfo[nicknamecol] != pinfo[firstnamecol]:
                    People(pinfo, thefields, firstname = pinfo[nicknamecol])

    @classmethod
    def loadpeople(self, fn, debug=False):  # This version is for years with one adult per record
        self.debug = debug
        commonfields = ('household_id', 'address', 'address2', 'city', 'state', 'zip')
        personfields = ('email', 'title', 'firstname', 'lastname', 'nickname')
        firstnamecol = personfields.index('firstname')
        nicknamecol = personfields.index('nickname')
        emailcol = personfields.index('email')
        thefields = personfields + commonfields
        p1fields = [f for f in personfields] + list(commonfields)
        s = load_workbook(fn)
        sheet = s.active
        self.setlabels(getLabelsFromSheet(sheet))
        for inrow in sheet.iter_rows(min_row=2):
            # Clean up the values and put them in a dictionary indexed by column label
            values = [normalize(v.value) for v in inrow]
            row = {self.labels[i]: values[i] for i in range(len(values))}



            # Load each person, taking firstname synonyms into account
            altnames = False
            # Some years, the data has 'address1' in it; other times, it's 'address'.  Other code wants 'address', so let's force that.
            if 'address' not in row and 'address1' in row:
                row['address'] = row['address1']
            # If we don't have an 'address2', create it.
            if 'address2' not in row:
                row['address2'] = ''
            # The code wants 'household_id' but the datadump has 'account_id'.  Fix it.
            row['household_id'] = row['account_id']
            # Handle hyphenated last names and composite last names
            lastparts = [row['lastname']]
            if '-' in row['lastname']:
                lastparts.extend(row['lastname'].split('-'))
            if ' ' in row['lastname']:
                lastparts.extend(row['lastname'].split(' '))
            for item in lastparts:  # will include any parts of a hyphenated name
                row['lastname'] = item.strip()
                pinfo = [row[field] for field in p1fields]
                People(pinfo, thefields)
                if pinfo[nicknamecol] and pinfo[nicknamecol] != pinfo[firstnamecol]:
                    People(pinfo, thefields, firstname=pinfo[nicknamecol])

    @classmethod
    def find(self, id):
        try:
            return self.people[id]
        except KeyError:
            print("Could not find %s" % id)
            return False
            
    @classmethod
    def findbyname(self, key):        
        try:
            if len(self.people[key]) > 1:
                print("%s has multiple entries; use ID!" % key)
            return self.people[key][0]
        except KeyError:
            # print("Could not find %s" % key)
            return False
            
    def getaddr(self):
        return "%s, %s, %s  %s" % (self.streetaddress, self.city, self.state, self.postalcode)
        
    def handlenickname(self):
        return
        """Process names where the roster has a formal name but we want to use an informal name"""
        if self.nickname and self.firstname != self.nickname:
            self.displayname = self.nickname + ' ' + self.lastname


    @classmethod
    def add_dummy(self, firstname=None, lastname=None, email=None):
        self.dummycount += 1
        labels = ['nickname', 'household_id', 'title', 'address', 'city', 'state', 'zip', 'email', 'address2']
        row = ['', f'dummy{self.dummycount}', '', '20 Cherryblossom', 'Los Gatos', 'CA', '95032', email, '']
        return People(row, labels, firstname, lastname)
	
    def __init__(self, row, labels, firstname=None, lastname=None):
        for x in range(len(row)):
            setattr(self, labels[x], row[x])

        # Handle special processing for names if needed:
        if firstname:
            self.firstname = firstname
        if lastname:
            self.lastname = lastname

        # Let's try to normalize the street address, at least for things like Dr/Dr./Drive
        sa = []
        for w in self.address.split():
            wc = w.replace('.','')
            wc = wc.replace(',','')
            wc = wc.lower()
            if wc in self.synonyms:
                sa.append(self.synonyms[wc] + (',' if ',' in w else ''))
            else:
                sa.append(w)
        self.address = ' '.join(sa)
        
        self.key = self.firstname + ' ' + self.lastname
        self.displayname = self.key
        self.handlenickname()
        
        self.internalcontactid = self.getlinenumber()
        self.people[self.internalcontactid] = self
    
        
        if self.debug and self.key in self.people:
            if not firstname and not lastname:
                print("Duplicate: %s" % self.key)
                other = self.people[self.key]
                for one in other:
                    print("  Old: ID=%5s, address=%s" % (one.internalcontactid, one.getaddr()))
                print("  New: ID=%5s, address=%s" % (self.internalcontactid, self.getaddr()))
            self.people[self.key].append(self)
        else:
            self.people[self.key] = [self]
 
    def __rexpr__(self):
        return "display = '%s', first = '%s', last = '%s', address = '%s', email = '%s', sendpaper = %s" % (self.displayname, self.firstname, self.lastname, self.streetaddress, self.email, self.sendpaper)
        

if __name__ == "__main__":
    from parms import Parms
    import os
    import sqlite3
    dbname = 'people.db'
    parms = Parms()
    os.chdir(parms.datadir)
    People.loadpeople(parms.roster)

    # Delete old database if it exists
    try:
        os.remove(dbname)
    except FileNotFoundError:
        pass

    conn = sqlite3.connect(dbname)
    cur = conn.cursor()
    cur.execute('''CREATE TABLE people
                (key text, household_id text, displayname text, firstname text, lastname text, 
                address text, address2 text, city text, state text, zip text, email text)''')

    datalist = []
    for pnum in People.people:
        person = People.people[pnum]
        if isinstance(person, People):
            datalist.append((pnum, person.household_id, person.displayname, person.firstname, person.lastname,
                             person.address, person.address2, person.city, person.state, person.zip, person.email))
        elif isinstance(person, list):
            for each in person:
                datalist.append((each.key, each.household_id, each.displayname, each.firstname, each.lastname,
                                 each.address, each.address2, each.city, each.state, each.zip, each.email))
        else:
            print(f'{type(person)} is not People or list)\n{person}')
        #print(person)
    cur.executemany('''INSERT INTO people VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', datalist)
    conn.commit()



